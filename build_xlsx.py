# -*- coding: utf-8 -*-
"""Generate roadmap.xlsx from tasks_data.TASKS"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from tasks_data import TASKS, RANKS, LEVEL_BASE, LEVEL_STEP

OUT = "/home/Fedora/red-team-tracker/roadmap.xlsx"

# ---- palette ----
DARK   = "0D1117"
PANEL  = "161B22"
ACCENT = "1F6FEB"
GREEN  = "1A7F37"
GREENL = "C6EFCE"
AMBER  = "9A6700"
AMBERL = "FFF3CD"
GREY   = "6E7681"
WHITE  = "FFFFFF"
HEADER = "21262D"

thin = Side(style="thin", color="30363D")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ============================================================= ROADMAP SHEET
ws = wb.active
ws.title = "Roadmap"

headers = ["ID", "Phase", "Track", "Category", "Task", "Difficulty", "XP",
           "Status", "Earned XP", "Date Done", "Notes"]
ws.append(headers)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for i, (tid, phase, track, cat, title, diff, xp) in enumerate(TASKS, start=2):
    diff_stars = "★" * diff
    ws.append([tid, phase, track, cat, title, diff_stars, xp,
               "Not Started", "", "", ""])
    # Earned XP formula: if Status == Done, XP else 0
    ws.cell(row=i, column=9).value = f'=IF(H{i}="Done",G{i},0)'
    for c in range(1, 12):
        cell = ws.cell(row=i, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="center",
                                   wrap_text=(c == 5),
                                   horizontal="center" if c in (1,2,3,4,6,7,8,9,10) else "left")
    ws.cell(row=i, column=1).font = Font(bold=True, color=ACCENT)

nrows = len(TASKS) + 1

# Status dropdown
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"H2:H{nrows}")

# Conditional formatting: Done -> green, In Progress -> amber
done_fill = PatternFill("solid", fgColor=GREENL)
prog_fill = PatternFill("solid", fgColor=AMBERL)
ws.conditional_formatting.add(f"A2:K{nrows}",
    FormulaRule(formula=['$H2="Done"'], fill=done_fill))
ws.conditional_formatting.add(f"A2:K{nrows}",
    FormulaRule(formula=['$H2="In Progress"'], fill=prog_fill))

# column widths
widths = [9, 10, 12, 12, 70, 11, 7, 13, 11, 13, 30]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{nrows}"

# ============================================================= DASHBOARD SHEET
db = wb.create_sheet("Dashboard", 0)
db.sheet_view.showGridLines = False

db.column_dimensions["A"].width = 3
db.column_dimensions["B"].width = 30
db.column_dimensions["C"].width = 18
db.column_dimensions["D"].width = 18
db.column_dimensions["E"].width = 18
db.column_dimensions["F"].width = 18

def title_cell(ref, text, size=20):
    db[ref] = text
    db[ref].font = Font(bold=True, size=size, color=ACCENT)

title_cell("B2", "RED TEAM MASTERY  —  PROGRESS DASHBOARD", 18)
db["B3"] = "From intermediate pentester to elite adversary-emulation operator"
db["B3"].font = Font(italic=True, color=GREY, size=10)

TOTAL_XP = sum(t[6] for t in TASKS)
TOTAL_TASKS = len(TASKS)

# KPI cards
def kpi(cell_label, cell_val, label, formula, val_color=ACCENT):
    db[cell_label] = label
    db[cell_label].font = Font(bold=True, color=GREY, size=10)
    db[cell_val] = formula
    db[cell_val].font = Font(bold=True, size=22, color=val_color)

kpi("B5", "B6", "TASKS DONE", f'=COUNTIF(Roadmap!H2:H{nrows},"Done")&"  /  {TOTAL_TASKS}"', GREEN)
kpi("C5", "C6", "XP EARNED",  f'=SUM(Roadmap!I2:I{nrows})', ACCENT)
kpi("D5", "D6", "TOTAL XP",   f'={TOTAL_XP}', GREY)
kpi("E5", "E6", "% COMPLETE", f'=TEXT(SUM(Roadmap!I2:I{nrows})/{TOTAL_XP},"0%")', AMBER)

# Level + rank (level = number of thresholds passed; cumulative curve)
# Build cumulative XP thresholds for levels in a helper block, then LOOKUP.
db["B8"] = "LEVEL"
db["B8"].font = Font(bold=True, color=GREY, size=10)
db["B9"] = f'=LOOKUP(SUM(Roadmap!I2:I{nrows}),LevelXP,LevelNo)'
db["B9"].font = Font(bold=True, size=28, color="D29922")

db["C8"] = "RANK"
db["C8"].font = Font(bold=True, color=GREY, size=10)
db["C9"] = '=LOOKUP(B9,RankLvl,RankName)'
db["C9"].font = Font(bold=True, size=16, color="D29922")
db.merge_cells("C9:F9")

# Per-phase progress table
db["B12"] = "PROGRESS BY PHASE"
db["B12"].font = Font(bold=True, size=12, color=WHITE)
db["B12"].fill = PatternFill("solid", fgColor=HEADER)
db.merge_cells("B12:F12")

ph_headers = ["Phase", "Tasks Done", "Task Count", "XP Earned", "XP Total"]
for c, h in enumerate(ph_headers, 2):
    cell = db.cell(row=13, column=c)
    cell.value = h
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=HEADER)
    cell.alignment = Alignment(horizontal="center")

phases = ["Phase 1", "Phase 2", "Phase 3", "Phase 4", "Phase 5", "Phase 6", "Tracks", "Capstone"]
r = 14
for ph in phases:
    db.cell(row=r, column=2, value=ph).font = Font(bold=True)
    db.cell(row=r, column=3, value=f'=COUNTIFS(Roadmap!B2:B{nrows},B{r},Roadmap!H2:H{nrows},"Done")')
    db.cell(row=r, column=4, value=f'=COUNTIF(Roadmap!B2:B{nrows},B{r})')
    db.cell(row=r, column=5, value=f'=SUMIFS(Roadmap!I2:I{nrows},Roadmap!B2:B{nrows},B{r})')
    db.cell(row=r, column=6, value=f'=SUMIFS(Roadmap!G2:G{nrows},Roadmap!B2:B{nrows},B{r})')
    for c in range(2, 7):
        db.cell(row=r, column=c).border = border
        if c > 2:
            db.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    r += 1

# Per-track table
db["B24"] = "PROGRESS BY TRACK"
db["B24"].font = Font(bold=True, size=12, color=WHITE)
db["B24"].fill = PatternFill("solid", fgColor=HEADER)
db.merge_cells("B24:F24")
for c, h in enumerate(["Track", "Tasks Done", "Task Count", "XP Earned", "XP Total"], 2):
    cell = db.cell(row=25, column=c)
    cell.value = h
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=HEADER)
    cell.alignment = Alignment(horizontal="center")

tracks = ["Core", "Technical", "Tradecraft", "Reporting", "Detection", "Purple"]
r = 26
for tk in tracks:
    db.cell(row=r, column=2, value=tk).font = Font(bold=True)
    db.cell(row=r, column=3, value=f'=COUNTIFS(Roadmap!C2:C{nrows},B{r},Roadmap!H2:H{nrows},"Done")')
    db.cell(row=r, column=4, value=f'=COUNTIF(Roadmap!C2:C{nrows},B{r})')
    db.cell(row=r, column=5, value=f'=SUMIFS(Roadmap!I2:I{nrows},Roadmap!C2:C{nrows},B{r})')
    db.cell(row=r, column=6, value=f'=SUMIFS(Roadmap!G2:G{nrows},Roadmap!C2:C{nrows},B{r})')
    for c in range(2, 7):
        db.cell(row=r, column=c).border = border
        if c > 2:
            db.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    r += 1

db["B33"] = "How to use: mark Status = 'Done' on the Roadmap tab. Earned XP, level, rank and all charts update automatically."
db["B33"].font = Font(italic=True, color=GREY, size=9)
db.merge_cells("B33:F33")

# ============================================================= LEVELS HELPER SHEET
lv = wb.create_sheet("Levels")
lv.sheet_state = "hidden"
lv["A1"] = "LevelNo"; lv["B1"] = "CumulativeXP"; lv["C1"] = "RankLvl"; lv["D1"] = "RankName"
# cumulative XP curve
cum = 0
level_rows = []
for L in range(1, 41):
    if L == 1:
        cum = 0
    else:
        cum += LEVEL_BASE + LEVEL_STEP * (L - 1)
    lv.cell(row=L+1, column=1, value=L)
    lv.cell(row=L+1, column=2, value=cum)
    level_rows.append((L, cum))
# ranks
for i, (rl, rn) in enumerate(RANKS, start=2):
    lv.cell(row=i, column=3, value=rl)
    lv.cell(row=i, column=4, value=rn)

# Named ranges for the LOOKUP() calls (LOOKUP needs ascending vectors)
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names.add(DefinedName("LevelXP", attr_text=f"Levels!$B$2:$B${len(level_rows)+1}"))
wb.defined_names.add(DefinedName("LevelNo", attr_text=f"Levels!$A$2:$A${len(level_rows)+1}"))
wb.defined_names.add(DefinedName("RankLvl", attr_text=f"Levels!$C$2:$C${len(RANKS)+1}"))
wb.defined_names.add(DefinedName("RankName", attr_text=f"Levels!$D$2:$D${len(RANKS)+1}"))

# ============================================================= LEGEND SHEET
lg = wb.create_sheet("Legend")
lg.sheet_view.showGridLines = False
lg.column_dimensions["A"].width = 3
lg.column_dimensions["B"].width = 22
lg.column_dimensions["C"].width = 80
lg["B2"] = "LEGEND & LEVELING"
lg["B2"].font = Font(bold=True, size=16, color=ACCENT)

rows = [
    ("Categories", ""),
    ("Learn", "Study/understand a concept or technique family"),
    ("Exercise", "Hands-on drill in your own lab"),
    ("Lab", "Full lab / range / Pro Lab environment"),
    ("Assessment", "Gate you must pass to advance (proves the skill)"),
    ("Milestone", "Certification or major checkpoint"),
    ("Track", "Ongoing parallel-track work (tradecraft, reporting, etc.)"),
    ("Capstone", "Independent end-to-end engagement simulation"),
    ("", ""),
    ("Tracks", ""),
    ("Core", "Main technical attack-chain progression"),
    ("Technical", "Offensive dev, exploit dev, maldev depth"),
    ("Tradecraft", "OPSEC, C2 infra, attribution discipline"),
    ("Reporting", "Executive + technical writing, engagement leadership"),
    ("Detection", "Detection-aware red teaming, Sigma/telemetry"),
    ("Purple", "Purple-team collaboration cycles"),
    ("", ""),
    ("Difficulty", "1 star (foundational) to 5 stars (elite)"),
    ("Leveling", f"Level up needs {LEVEL_BASE} + {LEVEL_STEP} x (level) XP each step; ranks unlock at levels 1/5/10/15/20/25/30"),
]
r = 4
for k, v in rows:
    if v == "" and k and not k.startswith(" "):
        lg.cell(row=r, column=2, value=k).font = Font(bold=True, size=12, color="D29922")
    else:
        lg.cell(row=r, column=2, value=k).font = Font(bold=True, color=WHITE if k else GREY)
        lg.cell(row=r, column=3, value=v).font = Font(color="C9D1D9")
    r += 1

wb.save(OUT)
print("Wrote", OUT, "with", TOTAL_TASKS, "tasks, total XP", TOTAL_XP)
